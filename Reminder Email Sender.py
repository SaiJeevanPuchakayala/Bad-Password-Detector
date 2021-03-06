#pip install openpyxl
#pip install smtplib
import openpyxl
import smtplib


# Loading the Weak Password Users.xlsx spreadsheet.
Wpasswrdsheet = openpyxl.load_workbook('Weak Password Users.xlsx')  
# Selecting sheet1 from Weak Password Users.xlsx spreadsheet.
sheet = Wpasswrdsheet["Sheet1"] 
# Assigning last column of sheet1.
lastCol = sheet.max_column 



# Dictionary to temporarily store weak passwords users name and email id.
Wpasswrdusers = {} 
for r in range(2, sheet.max_row + 1):
    strength = sheet.cell(row=r, column=lastCol).value
    if  strength!= 'Strong':
        user_name = sheet.cell(row=r, column=1).value
        email_id = sheet.cell(row=r, column=2).value
        Wpasswrdusers[user_name] = email_id




# Sends emails based on Password Strength status in spreadsheet.

 #Enter your gmail id here
senderemail = "#"          
#Enter password of your gmail
password = input(str("Please enter your password: ")) 
#Here the SMTP server is established
server = smtplib.SMTP('smtp.gmail.com',587)   
#Here the server is secured        
server.starttls()                                    
 #Here your gmail login takes place
server.login(senderemail,password)                   
print("****************Login Successful*********************")


for user_name, email_id  in Wpasswrdusers.items():
    #Enter your email subject here
    SUBJECT = "Weak Login Password."                         
    #Enter the actual email content here
    TEXT = "\nDear %s,\n\n\n\nRecords show that you are using weak password for \n\n Example.com. \n\n\nPlease make sure that you change your password immeadiately. \n\nThank you! \n\n\nTeam Example" %(user_name)

    #your message being structured here.
    message = 'Subject: {}\n\n{}'.format(SUBJECT, TEXT)
    print('Sending email to %s...' % email_id)
    
    #Here Email will be composed and sent to reciever gmail.
    sendmail_Status = server.sendmail(senderemail, email_id,message)

    if sendmail_Status != {}:
           print('There was a problem sending email to %s: %s' % (email_id,
           sendmail_Status))
server.quit()